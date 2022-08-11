import openpyxl

def get_repo_dict(file_path):
    MAX_ROWS=10000

    wb = openpyxl.load_workbook(file_path)
    ws = wb["Sheet1"]

    repo_dict = {}

    for num in range(MAX_ROWS):
        repository_id = ws.cell(row=num + 2, column=1).value
        if repository_id == None:
            break
        pull_request_id = ws.cell(row=num + 2, column=2).value

        if repository_id in repo_dict.keys():
            repo_dict[repository_id].append(pull_request_id)
        else:
            list = [pull_request_id]
            repo_dict[repository_id] = list

    return repo_dict

